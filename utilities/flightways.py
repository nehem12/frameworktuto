import inspect
import logging
import unittest
from openpyxl import load_workbook

class Flight_options(unittest.TestCase):
    def flways(self, liste, optionn):
        for m in liste:
            self.assertEqual(m.text, optionn)
            if m.text == optionn:
                print("Test passed")
            else:
                print("Test failed")

    def logss(logLevel=logging.DEBUG):
        logname = inspect.stack()[1][3]
        logger = logging.getLogger(logname)
        logger.setLevel(logLevel)
        log_file = logging.FileHandler("frameworklogs.log", mode='a')
        log_file.flush()
        log_formatter = logging.Formatter('%(levelname)s -%(name)s - %(asctime)s - %(message)s')
        log_file.setFormatter(log_formatter)
        logger.addHandler(log_file)
        return logger

    def dataexcel(fname,sheetname):
        datalist = []
        workobj = load_workbook(filename=fname)
        sheetobj = workobj[sheetname]
        rowcount = sheetobj.max_row
        columncount = sheetobj.max_column

        for i in range (2,rowcount+1):
            row = []
            for j in range (1,columncount+1):
                row.append(sheetobj.cell(row=i,column=j).value)
            datalist.append(row)
        return datalist



